"""
Report Export Utilities for PDF and Excel generation
"""
import io
from datetime import datetime
from typing import Dict, List
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ReportExporter:
    """Base class for report export functionality"""
    
    @staticmethod
    def get_status_color(status: str, format_type: str = 'hex'):
        """Get color for status with proper formatting"""
        colors_map = {
            'in_progress': {
                'hex': '#FEF3C7',  # yellow-100
                'rgb': (254, 243, 199),
                'reportlab': colors.Color(254/255, 243/255, 199/255)
            },
            'Pending with Presales': {
                'hex': '#E5E7EB',  # gray-200
                'rgb': (229, 231, 235),
                'reportlab': colors.Color(229/255, 231/255, 235/255)
            },
            'Pending review': {
                'hex': '#E9D5FF',  # purple-200
                'rgb': (233, 213, 255),
                'reportlab': colors.Color(233/255, 213/255, 255/255)
            },
            'Pending approval': {
                'hex': '#FED7AA',  # orange-200
                'rgb': (254, 215, 170),
                'reportlab': colors.Color(254/255, 215/255, 170/255)
            },
            'Closed Request': {
                'hex': '#DCFCE7',  # green-100
                'rgb': (220, 252, 231),
                'reportlab': colors.Color(220/255, 252/255, 231/255)
            }
        }
        
        default_color = {
            'hex': '#FFFFFF',
            'rgb': (255, 255, 255),
            'reportlab': colors.white
        }
        
        return colors_map.get(status, default_color).get(format_type, default_color[format_type])
    
    @staticmethod
    def is_overdue(request: Dict) -> bool:
        """Check if a request is overdue (including closed requests that exceeded target)"""
        return (request.get('target_days') and 
                request.get('duration_days', 0) > request.get('target_days', 0))


class PDFExporter(ReportExporter):
    """PDF export functionality"""
    
    def __init__(self):
        self.styles = getSampleStyleSheet()
        self.title_style = ParagraphStyle(
            'CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=TA_CENTER
        )
        self.table_cell_style = ParagraphStyle(
            'TableCell',
            parent=self.styles['Normal'],
            fontSize=8,
            alignment=TA_CENTER,
            spaceAfter=0,
            spaceBefore=0,
            leftIndent=0,
            rightIndent=0
        )
        self.table_header_style = ParagraphStyle(
            'TableHeader',
            parent=self.styles['Normal'],
            fontSize=7,
            alignment=TA_CENTER,
            spaceAfter=0,
            spaceBefore=0,
            leftIndent=1,
            rightIndent=1,
            fontName='Helvetica-Bold'
        )
        
    def create_report_pdf(self, report_data: Dict, report_type: str, period: str) -> io.BytesIO:
        """Create PDF report with color coding"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=0.5*inch, leftMargin=0.5*inch, rightMargin=0.5*inch)
        story = []
        
        # Header with logo and team info
        header_style = ParagraphStyle(
            'HeaderStyle',
            parent=self.styles['Normal'],
            fontSize=14,
            spaceAfter=10,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        # Company header
        story.append(Paragraph("GBB Solution Design Team", header_style))
        
        # Title with date
        title = f"{report_type.title()} Report - {period}"
        story.append(Paragraph(title, self.title_style))
        story.append(Spacer(1, 20))
        
        # Summary metrics with detailed status breakdown
        metrics_data = [
            ['Metric', 'Count'],
            ['Created', str(report_data.get('created', 0))],
            ['Completed', str(report_data.get('completed', 0))],
            ['In Progress', str(report_data.get('in_progress', 0))],
            ['Overdue', str(report_data.get('overdue', 0))]
        ]
        
        # Calculate status breakdown from requests data
        if 'requests' in report_data and report_data['requests']:
            status_counts = {}
            for req in report_data['requests']:
                status = req.get('status', 'Unknown')
                status_counts[status] = status_counts.get(status, 0) + 1
            
            metrics_data.extend([
                ['', ''],  # Empty row separator
                ['Status Breakdown', ''],
                ['In Progress', str(status_counts.get('in_progress', 0))],
                ['Pending with Presales', str(status_counts.get('Pending with Presales', 0))],
                ['Pending Review', str(status_counts.get('Pending review', 0))],
                ['Pending Approval', str(status_counts.get('Pending approval', 0))],
                ['Closed Requests', str(status_counts.get('Closed Request', 0))]
            ])
        
        metrics_table = Table(metrics_data, colWidths=[2*inch, 1*inch])
        metrics_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(Paragraph("Summary Metrics", self.styles['Heading2']))
        story.append(metrics_table)
        story.append(Spacer(1, 20))
        
        # Requests table
        if 'requests' in report_data and report_data['requests']:
            story.append(Paragraph("Request Details", self.styles['Heading2']))
            
            # Table headers with specified column order - wrapped as Paragraphs for proper text wrapping
            headers = [
                Paragraph('S/N', self.table_header_style),
                Paragraph('Customer', self.table_header_style),
                Paragraph('Description', self.table_header_style),
                Paragraph('BOQ-Cost (NGN)', self.table_header_style),
                Paragraph('BM (Name)', self.table_header_style),
                Paragraph('Date Request Received', self.table_header_style),
                Paragraph('Target (working days)', self.table_header_style),
                Paragraph('Date Sent Out (Date sent to BD/RDIS/EBG)', self.table_header_style),
                Paragraph('Duration (Working days)', self.table_header_style),
                Paragraph('Team Member Involved', self.table_header_style),
                Paragraph('Comment', self.table_header_style)
            ]
            table_data = [headers]
            
            # Table rows with color coding
            for index, req in enumerate(report_data['requests'], 1):
                comment = req.get('comment', '') or ''
                boq_cost = f"NGN {req.get('boq_cost', 0):,.2f}" if req.get('boq_cost') else 'N/A'
                sent_out_date = req.get('sent_out_date', '') if req.get('sent_out_date') else 'N/A'
                
                # Create Paragraph objects for text wrapping
                row = [
                    Paragraph(str(index), self.table_cell_style),  # Sequential numbering
                    Paragraph(req.get('customer_name', ''), self.table_cell_style),
                    Paragraph(req.get('description', ''), self.table_cell_style),
                    Paragraph(boq_cost, self.table_cell_style),
                    Paragraph(req.get('requester_name', ''), self.table_cell_style),
                    Paragraph(req.get('date_request_received', ''), self.table_cell_style),
                    Paragraph(f"{req.get('target_days', 'N/A')}" if req.get('target_days') else 'N/A', self.table_cell_style),
                    Paragraph(sent_out_date, self.table_cell_style),
                    Paragraph(f"{req.get('duration_days', 0)}", self.table_cell_style),
                    Paragraph(req.get('team_member_involved', ''), self.table_cell_style),
                    Paragraph(comment, self.table_cell_style)
                ]
                table_data.append(row)
            
            requests_table = Table(table_data, colWidths=[0.3*inch, 0.8*inch, 1.0*inch, 0.7*inch, 0.8*inch, 
                                                         0.8*inch, 0.5*inch, 0.8*inch, 0.5*inch, 0.8*inch, 0.9*inch], 
                                  repeatRows=1)
            
            # Base table style
            table_style = [
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 7),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
                ('LEFTPADDING', (0, 0), (-1, -1), 2),
                ('RIGHTPADDING', (0, 0), (-1, -1), 2),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 0), (-1, 0), [colors.grey]),
                ('WORDWRAP', (0, 0), (-1, -1), True)
            ]
            
            # Add color coding for each row
            for i, req in enumerate(report_data['requests'], 1):
                status_color = self.get_status_color(req.get('status', ''), 'reportlab')
                table_style.append(('BACKGROUND', (0, i), (-1, i), status_color))
                
                # Red background for overdue requests - only duration cell (column 8)
                if self.is_overdue(req):
                    table_style.append(('BACKGROUND', (8, i), (8, i), colors.Color(254/255, 202/255, 202/255)))  # red-200
            
            requests_table.setStyle(TableStyle(table_style))
            story.append(requests_table)
            
            # Color guide section
            story.append(Spacer(1, 20))
            story.append(Paragraph("Color Guide", self.styles['Heading2']))
            
            legend_data = [
                ['Status', 'Color'],
                ['Closed Request', ''],
                ['Pending with Presales', ''],
                ['Pending review', ''],
                ['Pending approval', ''],
                ['In Progress', ''],
                ['Overdue (Duration)', '']
            ]
            
            legend_table = Table(legend_data, colWidths=[2*inch, 1*inch])
            
            legend_style = [
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                # Status colors
                ('BACKGROUND', (1, 1), (1, 1), self.get_status_color('Closed Request', 'reportlab')),
                ('BACKGROUND', (1, 2), (1, 2), self.get_status_color('Pending with Presales', 'reportlab')),
                ('BACKGROUND', (1, 3), (1, 3), self.get_status_color('Pending review', 'reportlab')),
                ('BACKGROUND', (1, 4), (1, 4), self.get_status_color('Pending approval', 'reportlab')),
                ('BACKGROUND', (1, 5), (1, 5), self.get_status_color('in_progress', 'reportlab')),
                ('BACKGROUND', (1, 6), (1, 6), colors.Color(254/255, 202/255, 202/255))  # red-200 for overdue
            ]
            
            legend_table.setStyle(TableStyle(legend_style))
            story.append(legend_table)
        
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        return buffer


class ExcelExporter(ReportExporter):
    """Excel export functionality"""
    
    def create_report_excel(self, report_data: Dict, report_type: str, period: str) -> io.BytesIO:
        """Create Excel report with color coding"""
        wb = Workbook()
        ws = wb.active
        ws.title = f"{report_type.title()} Report"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Header with team info
        ws.merge_cells('A1:N1')
        ws['A1'] = "GBB Solution Design Team"
        ws['A1'].font = Font(bold=True, size=18, color="366092")
        ws['A1'].alignment = center_alignment
        
        # Title with date
        ws.merge_cells('A2:N2')
        ws['A2'] = f"{report_type.title()} Report - {period}"
        ws['A2'].font = Font(bold=True, size=16)
        ws['A2'].alignment = center_alignment
        
        # Summary metrics
        row = 4
        ws[f'A{row}'] = "Summary Metrics"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        
        row += 1
        metrics = [
            ('Created', report_data.get('created', 0)),
            ('Completed', report_data.get('completed', 0)),
            ('In Progress', report_data.get('in_progress', 0)),
            ('Overdue', report_data.get('overdue', 0))
        ]
        
        for metric, value in metrics:
            ws[f'A{row}'] = metric
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border
            row += 1
        
        # Calculate status breakdown from requests data
        if 'requests' in report_data and report_data['requests']:
            status_counts = {}
            for req in report_data['requests']:
                status = req.get('status', 'Unknown')
                status_counts[status] = status_counts.get(status, 0) + 1
            
            row += 1  # Empty row separator
            
            # Status breakdown header
            ws[f'A{row}'] = "Status Breakdown"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
            
            status_metrics = [
                ('In Progress', status_counts.get('in_progress', 0)),
                ('Pending with Presales', status_counts.get('Pending with Presales', 0)),
                ('Pending Review', status_counts.get('Pending review', 0)),
                ('Pending Approval', status_counts.get('Pending approval', 0)),
                ('Closed Requests', status_counts.get('Closed Request', 0))
            ]
            
            for metric, value in status_metrics:
                ws[f'A{row}'] = metric
                ws[f'B{row}'] = value
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'A{row}'].border = border
                ws[f'B{row}'].border = border
                row += 1
        
        # Requests section
        if 'requests' in report_data and report_data['requests']:
            row += 2
            ws[f'A{row}'] = "Request Details"
            ws[f'A{row}'].font = Font(bold=True, size=14)
            
            row += 1
            headers = ['S/N', 'Customer', 'Description', 'BOQ-Cost', 'BM (Name)', 
                      'Date Request Received', 'Target (working days)', 
                      'Date Sent Out (Date sent to BD/RDIS/EBG)', 'Duration (Working days)', 'Team Member Involved', 'Comment']
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = border
            
            # Data rows with color coding
            for index, req in enumerate(report_data['requests'], 1):
                row += 1
                comment = req.get('comment', '') or ''
                boq_cost = f"NGN {req.get('boq_cost', 0):,.2f}" if req.get('boq_cost') else 'N/A'
                sent_out_date = req.get('sent_out_date', '') if req.get('sent_out_date') else 'N/A'
                
                data = [
                    index,  # Sequential numbering
                    req.get('customer_name', ''),
                    req.get('description', ''),
                    boq_cost,
                    req.get('requester_name', ''),
                    req.get('date_request_received', ''),
                    req.get('target_days', '') if req.get('target_days') else 'N/A',
                    sent_out_date,
                    req.get('duration_days', 0),
                    req.get('team_member_involved', ''),
                    comment
                ]
                
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = border
                    cell.alignment = center_alignment if col in [1, 7, 9] else Alignment(horizontal="left", vertical="center")
                    
                    # Status color coding
                    status = req.get('status', '')
                    if status:
                        hex_color = self.get_status_color(status, 'hex').replace('#', '')
                        cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                    
                    # Overdue highlighting (red background) - only for duration column (column 9)
                    if self.is_overdue(req) and col == 9:
                        cell.fill = PatternFill(start_color="FECACA", end_color="FECACA", fill_type="solid")  # red-200
        
        # Color guide section
        if 'requests' in report_data and report_data['requests']:
            row += 3
            ws[f'A{row}'] = "Color Guide"
            ws[f'A{row}'].font = Font(bold=True, size=14)
            
            row += 1
            legend_headers = ['Status', 'Color']
            
            for col, header in enumerate(legend_headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = border
            
            # Legend data
            legend_items = [
                ('Closed Request', 'Closed Request'),
                ('Pending with Presales', 'Pending with Presales'),
                ('Pending review', 'Pending review'),
                ('Pending approval', 'Pending approval'),
                ('In Progress', 'in_progress'),
                ('Overdue (Duration)', 'overdue')
            ]
            
            for status_label, status_key in legend_items:
                row += 1
                # Status label
                cell = ws.cell(row=row, column=1, value=status_label)
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
                # Color cell
                cell = ws.cell(row=row, column=2, value="")
                cell.border = border
                if status_key == 'overdue':
                    cell.fill = PatternFill(start_color="FECACA", end_color="FECACA", fill_type="solid")  # red-200
                else:
                    hex_color = self.get_status_color(status_key, 'hex').replace('#', '')
                    cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
        
        
        # Auto-adjust column widths
        for col in range(1, 12):
            column_letter = get_column_letter(col)
            max_length = 0
            for row_cells in ws[f'{column_letter}1:{column_letter}{ws.max_row}']:
                for cell in row_cells:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 30)
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer